"""Dataset label exports for speech-enhancement scene captures."""

from __future__ import annotations

import csv
import json
import os
from pathlib import Path
from typing import Any

from .professional import array_geometry_sha256, array_metadata


LABEL_COLUMNS = [
    ("run_id", "运行目录编号", "本次采集运行目录名，用于跨运行汇总"),
    ("sample_id", "样本编号", "一次播录样本的稳定唯一编号"),
    ("supervision_pair_id", "监督配对编号", "同一行 mixture 与 target-only 的配对主键"),
    ("capture_type", "数据类型", "supervised_pair、target_only 或 interferer_only"),
    ("scene_id", "物理场景编号", "佩戴、麦杆、干扰源位置等固定的一次物理场景"),
    ("experiment_id", "实验编号", "一次固定物理条件实验的编号"),
    ("project_id", "项目编号", "用于跨运行汇总筛选的项目编号"),
    ("manual_label", "人工标签", "留给人工填写或修订的标签"),
    ("automatic_label", "自动标签", "由标签前缀和源文件名自动生成"),
    ("dataset_split", "数据集划分", "train、valid 或 test"),
    ("valid", "是否有效", "人工质检开关：是或否"),
    ("quality_flag", "质量标记", "自动质检结果"),
    ("quality_gate_enabled", "是否启用自动质量门限", "由项目模板统一设置，不要求现场人工评分"),
    ("quality_gate_passed", "自动质量门限是否通过", "仅在启用门限时影响监督可用状态"),
    ("mixture_consistency_residual_db_max", "混合可加性最大残差（dB）", "各麦克风中 mixture-(target+interferer) 的最大相对残差"),
    ("mixture_consistency_correlation_min", "混合可加性最低相关系数", "各麦克风中 mixture 与 target+interferer 的最低相关系数"),
    ("mixture_consistency_metrics_json", "混合可加性分通道指标", "自动质检指标 JSON；真实声学录音不要求精确相加"),
    ("supervision_quality_metrics", "监督质量指标文件", "相对运行目录的自动质检 JSON"),
    ("supervision_ready", "监督训练可用", "混合录音是否具有样本级对齐的仅目标标签"),
    ("capture_strategy", "采集策略", "single_stream_paired_sequence 表示同一声卡流内配对采集"),
    ("shared_hardware_clock", "输入输出共享硬件时钟", "输入和输出是否使用同一块可验证的双工声卡"),
    ("target_mixture_sample_aligned", "目标与混合样本对齐", "仅目标与混合的目标片段是否使用同一源、起点和长度"),
    ("interferer_mixture_sample_aligned", "干扰与混合样本对齐", "仅干扰与混合的干扰片段是否使用同一源、起点和长度"),
    ("supervision_contract", "监督配对约束", "监督配对所满足的采集与样本复用约束"),
    ("repetition", "重复编号", "同一物理条件下的重复编号"),
    ("pair_index", "配对序号", "本次文件夹配对计划中的一基序号"),
    ("pairing_strategy", "配对策略", "生成目标与干扰素材配对表的确定性算法版本"),
    ("pairing_seed", "配对随机种子", "相同素材与相同种子生成相同配对"),
    ("measurement_count", "计划测量次数", "本次随机清单要求完成的素材对数量"),
    ("target_source", "目标源文件", "目标干净语音源文件"),
    ("interferer_source", "干扰源文件", "干扰声音源文件"),
    ("target_source_sha256", "目标源 SHA256", "目标源文件内容哈希"),
    ("interferer_source_sha256", "干扰源 SHA256", "干扰源文件内容哈希"),
    ("duration_s", "时长（秒）", "本条播录的实际时长"),
    ("sample_rate_hz", "采样率（赫兹）", "录制采样率"),
    ("microphone_channels", "麦克风通道", "一基编号，逗号分隔"),
    ("target_output_channel", "目标输出通道", "目标声源所用声卡输出"),
    ("interferer_output_channel", "干扰输出通道", "干扰声源所用声卡输出"),
    ("target_level_dbfs", "目标电平（dBFS）", "目标数字播放峰值电平"),
    ("interferer_level_dbfs", "干扰电平（dBFS）", "干扰数字播放峰值电平"),
    ("room_id", "房间编号", "录音房间或声学环境编号"),
    ("artificial_head_id", "人工头编号", "佩戴耳机的人工头个体"),
    ("headset_model_id", "耳机型号", "耳机型号编号"),
    ("headset_unit_id", "耳机个体", "同型号下的物理耳机个体"),
    ("wearing_id", "佩戴编号", "一次独立重新佩戴的编号"),
    ("boom_pose_id", "麦杆姿态编号", "一次独立麦杆位置或姿态"),
    ("microphone_map_json", "麦克风映射", "录制通道到麦克风含义的 JSON 映射"),
    ("capture_profile", "采集级别", "standard/development/production"),
    ("task_type", "训练任务", "speech_enhancement/neural_beamforming/rir 等"),
    ("array_id", "阵列编号", "稳定的麦克风阵列或耳机阵列编号"),
    ("array_coordinate_system", "阵列坐标系", "阵列坐标原点和轴方向定义"),
    ("array_reference_point", "阵列参考点", "几何测量的参考点"),
    ("array_reference_channel", "阵列参考通道", "用于延时/相位参考的一基录制通道"),
    ("array_geometry_sha256", "阵列几何 SHA256", "阵列定义的稳定哈希，便于跨运行核对"),
    ("array_geometry_json", "阵列几何", "通道、坐标、极性和校准信息 JSON"),
    ("physical_capture_group_id", "物理采集分组", "房间、人工头、耳机个体、佩戴和麦杆姿态组合"),
    ("split_group_id", "数据划分分组", "同一分组不得同时出现在 train/valid/test"),
    ("speaker_id", "说话人编号", "目标语音说话人，用于防止数据划分泄漏"),
    ("utterance_id", "语句编号", "目标语句编号，用于防止数据划分泄漏"),
    ("noise_id", "干扰素材编号", "干扰录音或干扰事件编号"),
    ("noise_class", "干扰类别", "噪声/干扰的语义类别"),
    ("target_source_id", "目标声源编号", "人工嘴或目标扬声器编号"),
    ("target_position_id", "目标位置编号", "目标声源位置编号"),
    ("target_azimuth_deg", "目标方位角（度）", "目标声源相对人工头方位角"),
    ("target_elevation_deg", "目标俯仰角（度）", "目标声源相对人工头俯仰角"),
    ("target_height_m", "目标高度（米）", "目标声源实测高度"),
    ("target_distance_m", "目标距离（米）", "目标声源到参考点实测距离"),
    ("interferer_source_id", "干扰声源编号", "干扰扬声器或声源编号"),
    ("interferer_position_id", "干扰位置编号", "干扰声源位置编号"),
    ("interferer_azimuth_deg", "干扰方位角（度）", "干扰源相对人工头方位角"),
    ("interferer_elevation_deg", "干扰俯仰角（度）", "干扰源相对人工头俯仰角"),
    ("interferer_height_m", "干扰高度（米）", "干扰源实测高度"),
    ("interferer_distance_m", "干扰距离（米）", "干扰源到参考点实测距离"),
    ("ambient_recording", "环境底噪录音", "相对于本次运行目录的路径"),
    ("target_recording", "仅目标录音", "相对于本次运行目录的路径"),
    ("interferer_recording", "仅干扰录音", "相对于本次运行目录的路径"),
    ("mixture_recording", "混合录音", "相对于本次运行目录的路径"),
    ("target_playback", "目标播放参考", "两路输出矩阵，未使用通道为零"),
    ("interferer_playback", "干扰播放参考", "两路输出矩阵，未使用通道为零"),
    ("mixture_playback", "混合播放参考", "目标与干扰播放矩阵之和"),
    ("paired_sequence_recording", "配对序列完整录音", "一次连续声卡流的完整麦克风录音"),
    ("paired_sequence_playback", "配对序列完整播放", "一次连续声卡流的完整播放矩阵"),
    ("segment_layout", "片段边界文件", "各配对片段在连续声卡流中的起止采样点"),
    ("target_segment_start_sample", "目标片段起点", "target-only 在连续声卡流内的起始采样点"),
    ("mixture_segment_start_sample", "混合片段起点", "mixture 在连续声卡流内的起始采样点"),
    ("metrics_files", "指标文件", "各场景指标 JSON 路径"),
    ("metadata_json", "实验元数据", "人工头、佩戴、位置等 JSON"),
    ("notes", "备注", "可人工补充"),
]

CORE_LABEL_KEYS = [
    "run_id",
    "sample_id",
    "supervision_pair_id",
    "capture_type",
    "scene_id",
    "manual_label",
    "automatic_label",
    "dataset_split",
    "valid",
    "quality_flag",
    "quality_gate_passed",
    "mixture_consistency_residual_db_max",
    "mixture_consistency_correlation_min",
    "supervision_ready",
    "shared_hardware_clock",
    "target_mixture_sample_aligned",
    "repetition",
    "pair_index",
    "pairing_seed",
    "measurement_count",
    "duration_s",
    "notes",
]

EXPERIMENT_CONDITION_KEYS = [
    "sample_id",
    "project_id",
    "experiment_id",
    "room_id",
    "artificial_head_id",
    "headset_model_id",
    "headset_unit_id",
    "wearing_id",
    "boom_pose_id",
    "microphone_map_json",
    "capture_profile",
    "task_type",
    "array_id",
    "array_coordinate_system",
    "array_reference_point",
    "array_reference_channel",
    "array_geometry_sha256",
    "array_geometry_json",
    "physical_capture_group_id",
    "split_group_id",
    "speaker_id",
    "utterance_id",
    "noise_id",
    "noise_class",
    "target_source_id",
    "target_position_id",
    "target_azimuth_deg",
    "target_elevation_deg",
    "target_height_m",
    "target_distance_m",
    "interferer_source_id",
    "interferer_position_id",
    "interferer_azimuth_deg",
    "interferer_elevation_deg",
    "interferer_height_m",
    "interferer_distance_m",
]

SUPERVISED_PAIR_KEYS = [
    "supervision_pair_id",
    "sample_id",
    "scene_id",
    "dataset_split",
    "valid",
    "quality_flag",
    "supervision_ready",
    "quality_gate_enabled",
    "quality_gate_passed",
    "mixture_consistency_residual_db_max",
    "mixture_consistency_correlation_min",
    "supervision_quality_metrics",
    "shared_hardware_clock",
    "pair_index",
    "pairing_strategy",
    "pairing_seed",
    "measurement_count",
    "microphone_channels",
    "sample_rate_hz",
    "duration_s",
    "mixture_recording",
    "target_recording",
    "interferer_recording",
    "target_source_sha256",
    "interferer_source_sha256",
    "paired_sequence_recording",
    "segment_layout",
]

CAPTURE_PARAMETER_KEYS = [
    "sample_id",
    "capture_strategy",
    "pair_index",
    "pairing_strategy",
    "pairing_seed",
    "measurement_count",
    "sample_rate_hz",
    "microphone_channels",
    "target_output_channel",
    "interferer_output_channel",
    "target_level_dbfs",
    "interferer_level_dbfs",
]

FILE_INDEX_KEYS = [
    "sample_id",
    "target_source",
    "interferer_source",
    "ambient_recording",
    "target_recording",
    "interferer_recording",
    "mixture_recording",
    "target_playback",
    "interferer_playback",
    "mixture_playback",
    "paired_sequence_recording",
    "paired_sequence_playback",
    "segment_layout",
    "metrics_files",
    "metadata_json",
]


def flatten_experiment_metadata(metadata: dict[str, Any]) -> dict[str, Any]:
    """Expose training-relevant scene metadata as ordinary table columns."""
    target = metadata.get("target") if isinstance(metadata.get("target"), dict) else {}
    interferer = (
        metadata.get("interferer")
        if isinstance(metadata.get("interferer"), dict)
        else {}
    )
    array = array_metadata(metadata)
    array_channels = array.get("channels") if isinstance(array.get("channels"), list) else []
    microphones = {
        key.removeprefix("microphone_"): value
        for key, value in metadata.items()
        if key.startswith("microphone_") and value not in (None, "")
    }
    if array_channels:
        microphones = {
            str(channel.get("recording_channel")): channel.get("microphone_id", "")
            for channel in array_channels
            if isinstance(channel, dict) and channel.get("recording_channel") not in (None, "")
        }
    physical_parts = [
        str(metadata.get(key) or "")
        for key in (
            "room_id",
            "artificial_head_id",
            "headset_model_id",
            "headset_unit_id",
            "wearing_id",
            "boom_pose_id",
        )
    ]
    physical_capture_group_id = (
        "|".join(physical_parts) if any(physical_parts) else ""
    )
    return {
        "project_id": metadata.get("project_id", ""),
        "experiment_id": metadata.get("experiment_id", ""),
        "room_id": metadata.get("room_id", ""),
        "artificial_head_id": metadata.get("artificial_head_id", ""),
        "headset_model_id": metadata.get("headset_model_id", ""),
        "headset_unit_id": metadata.get("headset_unit_id", ""),
        "wearing_id": metadata.get("wearing_id", ""),
        "boom_pose_id": metadata.get("boom_pose_id", ""),
        "microphone_map_json": microphones,
        "capture_profile": metadata.get("capture_profile", "standard"),
        "task_type": metadata.get("task_type", metadata.get("experiment_type", "")),
        "array_id": array.get("array_id", ""),
        "array_coordinate_system": array.get("coordinate_system", ""),
        "array_reference_point": array.get("reference_point", ""),
        "array_reference_channel": array.get("reference_channel", ""),
        "array_geometry_sha256": array_geometry_sha256(metadata),
        "array_geometry_json": array,
        "physical_capture_group_id": physical_capture_group_id,
        "split_group_id": metadata.get("split_group_id", physical_capture_group_id),
        "speaker_id": metadata.get("speaker_id", ""),
        "utterance_id": metadata.get("utterance_id", ""),
        "noise_id": metadata.get("noise_id", ""),
        "noise_class": metadata.get("noise_class", ""),
        **_source_metadata("target", target),
        **_source_metadata("interferer", interferer),
    }


def _source_metadata(prefix: str, source: dict[str, Any]) -> dict[str, Any]:
    return {
        f"{prefix}_source_id": source.get("source_id", ""),
        f"{prefix}_position_id": source.get("position_id", ""),
        f"{prefix}_azimuth_deg": source.get("azimuth_deg", ""),
        f"{prefix}_elevation_deg": source.get("elevation_deg", ""),
        f"{prefix}_height_m": source.get("height_m", ""),
        f"{prefix}_distance_m": source.get("distance_m", ""),
    }


def write_label_files(
    run_dir: str | Path,
    rows: list[dict[str, Any]],
    metadata: dict[str, Any],
) -> dict[str, Path]:
    """Write JSONL, CSV and a formatted/editable XLSX workbook."""
    root = Path(run_dir)
    root.mkdir(parents=True, exist_ok=True)
    normalized = [
        {
            key: _cell_value(row.get(key, ""))
            for key, _header, _description in LABEL_COLUMNS
        }
        for row in rows
    ]

    jsonl_path = root / "labels.jsonl"
    with jsonl_path.open("w", encoding="utf-8") as handle:
        for row in normalized:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")

    csv_path = root / "labels.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=[key for key, _header, _description in LABEL_COLUMNS])
        writer.writeheader()
        writer.writerows(normalized)

    # This is a pairing table, not a pass-only training export. Failed rows
    # remain visible so the operator can audit them; the acceptance decision
    # is carried by supervision_ready and quality_flag.
    paired = [
        row
        for row in normalized
        if row.get("mixture_recording")
        and row.get("target_recording")
    ]
    supervised_jsonl_path = root / "supervised_pairs.jsonl"
    with supervised_jsonl_path.open("w", encoding="utf-8") as handle:
        for row in paired:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")
    supervised_csv_path = root / "supervised_pairs.csv"
    with supervised_csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=[key for key, _header, _description in LABEL_COLUMNS])
        writer.writeheader()
        writer.writerows(paired)

    xlsx_path = root / "labels.xlsx"
    _write_xlsx(xlsx_path, normalized, metadata)
    return {
        "jsonl": jsonl_path,
        "csv": csv_path,
        "xlsx": xlsx_path,
        "supervised_jsonl": supervised_jsonl_path,
        "supervised_csv": supervised_csv_path,
    }


def write_label_checkpoint(run_dir: str | Path, rows: list[dict[str, Any]]) -> dict[str, Path]:
    """Atomically checkpoint lightweight labels after each completed source pair."""
    root = Path(run_dir)
    root.mkdir(parents=True, exist_ok=True)
    normalized = [
        {key: _cell_value(row.get(key, "")) for key, _header, _description in LABEL_COLUMNS}
        for row in rows
    ]
    jsonl_path = root / "labels.partial.jsonl"
    csv_path = root / "labels.partial.csv"
    jsonl_tmp = jsonl_path.with_suffix(".jsonl.tmp")
    with jsonl_tmp.open("w", encoding="utf-8") as handle:
        for row in normalized:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")
    jsonl_tmp.replace(jsonl_path)
    csv_tmp = csv_path.with_suffix(".csv.tmp")
    with csv_tmp.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle, fieldnames=[key for key, _header, _description in LABEL_COLUMNS]
        )
        writer.writeheader()
        writer.writerows(normalized)
    csv_tmp.replace(csv_path)
    return {"jsonl": jsonl_path, "csv": csv_path}


def append_label_checkpoint(run_dir: str | Path, row: dict[str, Any]) -> Path:
    """Append one durable completed-pair label in O(1) time."""
    root = Path(run_dir)
    root.mkdir(parents=True, exist_ok=True)
    normalized = {
        key: _cell_value(row.get(key, "")) for key, _header, _description in LABEL_COLUMNS
    }
    path = root / "labels.partial.jsonl"
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(normalized, ensure_ascii=False) + "\n")
        handle.flush()
        os.fsync(handle.fileno())
    return path


def import_reviewed_labels(
    run_dir: str | Path, workbook_path: str | Path | None = None
) -> dict[str, Path]:
    """Import manual review columns from labels.xlsx without modifying raw labels."""
    root = Path(run_dir)
    source = Path(workbook_path) if workbook_path else root / "labels.xlsx"
    if not source.is_file():
        raise FileNotFoundError(f"找不到标签质检表：{source}")
    raw_path = root / "labels.jsonl"
    if not raw_path.is_file():
        raise FileNotFoundError(f"运行目录缺少 labels.jsonl：{root}")
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - installation failure
        raise RuntimeError("缺少 openpyxl，无法导入人工质检 Excel") from exc

    original = [
        json.loads(line)
        for line in raw_path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]
    by_id = {str(row.get("sample_id", "")): dict(row) for row in original}
    workbook = load_workbook(source, read_only=True, data_only=True)
    if "标签" not in workbook.sheetnames:
        raise ValueError("Excel 中缺少“标签”工作表")
    sheet = workbook["标签"]
    header_to_key = {header: key for key, header, _description in LABEL_COLUMNS}
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration as exc:
        raise ValueError("“标签”工作表为空") from exc
    keys = [header_to_key.get(str(value or ""), "") for value in headers]
    if "sample_id" not in keys:
        raise ValueError("“标签”工作表缺少“样本编号”列")
    review_keys = {"manual_label", "dataset_split", "valid", "notes"}
    seen: set[str] = set()
    for values in rows:
        record = {key: value for key, value in zip(keys, values) if key}
        sample_id = str(record.get("sample_id") or "").strip()
        if not sample_id:
            continue
        if sample_id not in by_id:
            raise ValueError(f"Excel 含有未知样本编号：{sample_id}")
        if sample_id in seen:
            raise ValueError(f"Excel 含有重复样本编号：{sample_id}")
        seen.add(sample_id)
        for key in review_keys:
            if key in record and record[key] is not None:
                by_id[sample_id][key] = str(record[key]).strip()
    if original and seen != set(by_id):
        missing = sorted(set(by_id) - seen)
        raise ValueError(
            "Excel 缺少原始标签中的样本，未导入以防误删：" + ", ".join(missing[:10])
        )
    reviewed = [by_id[str(row.get("sample_id", ""))] for row in original]
    for row in reviewed:
        if row.get("dataset_split") not in {"train", "valid", "test"}:
            raise ValueError(f"{row.get('sample_id')} 的数据集划分必须是 train/valid/test")
        if row.get("valid") not in {"是", "否"}:
            raise ValueError(f"{row.get('sample_id')} 的是否有效必须是“是”或“否”")

    reviewed_jsonl = root / "labels_reviewed.jsonl"
    reviewed_csv = root / "labels_reviewed.csv"
    with reviewed_jsonl.open("w", encoding="utf-8") as handle:
        for row in reviewed:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")
    with reviewed_csv.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=[key for key, _h, _d in LABEL_COLUMNS])
        writer.writeheader()
        writer.writerows(
            {key: row.get(key, "") for key, _h, _d in LABEL_COLUMNS} for row in reviewed
        )
    return {"jsonl": reviewed_jsonl, "csv": reviewed_csv}


def _cell_value(value: Any):
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return value


def _excel_column(index: int) -> str:
    value = index + 1
    letters = ""
    while value:
        value, remainder = divmod(value - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _write_xlsx(path: Path, rows: list[dict[str, Any]], metadata: dict[str, Any]) -> None:
    try:
        import xlsxwriter
    except Exception as exc:  # pragma: no cover - installation failure
        raise RuntimeError("缺少 XlsxWriter，无法生成 labels.xlsx") from exc

    workbook = xlsxwriter.Workbook(path)
    workbook.set_properties(
        {
            "title": "语音增强采集标签",
            "subject": "声学采集工具自动生成的数据集索引",
            "author": "声学采集工具",
            "comments": "人工标签、是否有效和备注列可在 Excel 中继续编辑。",
        }
    )
    header = workbook.add_format(
        {
            "bold": True,
            "font_color": "#FFFFFF",
            "bg_color": "#1F4E78",
            "align": "center",
            "valign": "vcenter",
            "border": 0,
        }
    )
    text_format = workbook.add_format({"valign": "top"})
    wrapped = workbook.add_format({"valign": "top", "text_wrap": True})
    number = workbook.add_format({"num_format": "0.000", "valign": "top"})
    section = workbook.add_format(
        {"bold": True, "font_color": "#FFFFFF", "bg_color": "#548235", "align": "left"}
    )
    note = workbook.add_format({"font_color": "#595959", "text_wrap": True, "valign": "top"})
    bad = workbook.add_format({"bg_color": "#F4CCCC", "font_color": "#9C0006"})
    warning = workbook.add_format({"bg_color": "#FFF2CC", "font_color": "#7F6000"})

    sheet = workbook.add_worksheet("标签")
    sheet.hide_gridlines(2)
    sheet.freeze_panes(1, 3)
    column_info = {key: (header_text, description) for key, header_text, description in LABEL_COLUMNS}
    keys = CORE_LABEL_KEYS
    headers = [column_info[key][0] for key in keys]
    for column, header_text in enumerate(headers):
        sheet.write(0, column, header_text, header)
    sheet.set_row(0, 30)

    for row_index, row in enumerate(rows, 1):
        for column, key in enumerate(keys):
            value = row.get(key, "")
            cell_format = number if key in {"duration_s", "target_level_dbfs", "interferer_level_dbfs"} else text_format
            if key == "notes":
                cell_format = wrapped
            sheet.write(row_index, column, value, cell_format)

    last_row = max(1, len(rows))
    last_column = len(keys) - 1
    split_column = keys.index("dataset_split")
    valid_column = keys.index("valid")
    quality_column = keys.index("quality_flag")
    sheet.autofilter(0, 0, last_row, last_column)
    sheet.data_validation(1, split_column, max(last_row, 1000), split_column, {"validate": "list", "source": ["train", "valid", "test"]})
    sheet.data_validation(1, valid_column, max(last_row, 1000), valid_column, {"validate": "list", "source": ["是", "否"]})
    sheet.conditional_format(1, valid_column, max(last_row, 1000), valid_column, {"type": "text", "criteria": "containing", "value": "否", "format": bad})
    sheet.conditional_format(1, quality_column, max(last_row, 1000), quality_column, {"type": "text", "criteria": "not containing", "value": "通过", "format": warning})

    widths = {
        0: 34,
        1: 18,
        2: 30,
        3: 12,
        4: 10,
        5: 18,
        6: 10,
        7: 12,
        8: 24,
    }
    for column in range(len(keys)):
        sheet.set_column(column, column, widths.get(column, 18))

    parameter_sheet = workbook.add_worksheet("采集参数")
    parameter_sheet.hide_gridlines(2)
    parameter_sheet.freeze_panes(1, 1)
    for column, key in enumerate(CAPTURE_PARAMETER_KEYS):
        parameter_sheet.write(0, column, column_info[key][0], header)
    parameter_sheet.set_row(0, 30)
    for row_index, row in enumerate(rows, 1):
        for column, key in enumerate(CAPTURE_PARAMETER_KEYS):
            cell_format = number if key in {"target_level_dbfs", "interferer_level_dbfs"} else text_format
            parameter_sheet.write(row_index, column, row.get(key, ""), cell_format)
    parameter_sheet.autofilter(0, 0, max(1, len(rows)), len(CAPTURE_PARAMETER_KEYS) - 1)
    parameter_sheet.set_column(0, 0, 34)
    parameter_sheet.set_column(1, len(CAPTURE_PARAMETER_KEYS) - 1, 18)

    condition_sheet = workbook.add_worksheet("实验条件")
    condition_sheet.hide_gridlines(2)
    condition_sheet.freeze_panes(1, 1)
    for column, key in enumerate(EXPERIMENT_CONDITION_KEYS):
        condition_sheet.write(0, column, column_info[key][0], header)
    for row_index, row in enumerate(rows, 1):
        for column, key in enumerate(EXPERIMENT_CONDITION_KEYS):
            condition_sheet.write(row_index, column, row.get(key, ""), wrapped)
    condition_sheet.autofilter(
        0, 0, max(1, len(rows)), len(EXPERIMENT_CONDITION_KEYS) - 1
    )
    condition_sheet.set_column(0, 0, 34)
    condition_sheet.set_column(1, len(EXPERIMENT_CONDITION_KEYS) - 1, 20)

    supervised_rows = [row for row in rows if row.get("supervision_ready") == "是"]
    supervised_sheet = workbook.add_worksheet("监督配对")
    supervised_sheet.hide_gridlines(2)
    supervised_sheet.freeze_panes(1, 2)
    for column, key in enumerate(SUPERVISED_PAIR_KEYS):
        supervised_sheet.write(0, column, column_info[key][0], header)
    for row_index, row in enumerate(supervised_rows, 1):
        for column, key in enumerate(SUPERVISED_PAIR_KEYS):
            supervised_sheet.write(row_index, column, row.get(key, ""), wrapped)
    supervised_sheet.autofilter(
        0, 0, max(1, len(supervised_rows)), len(SUPERVISED_PAIR_KEYS) - 1
    )
    supervised_sheet.set_column(0, 1, 34)
    supervised_sheet.set_column(2, len(SUPERVISED_PAIR_KEYS) - 1, 22)

    index_sheet = workbook.add_worksheet("文件索引")
    index_sheet.hide_gridlines(2)
    index_sheet.freeze_panes(1, 1)
    for column, key in enumerate(FILE_INDEX_KEYS):
        index_sheet.write(0, column, column_info[key][0], header)
    index_sheet.set_row(0, 30)
    for row_index, row in enumerate(rows, 1):
        for column, key in enumerate(FILE_INDEX_KEYS):
            index_sheet.write(row_index, column, row.get(key, ""), wrapped)
        index_sheet.set_row(row_index, 48)
    index_sheet.autofilter(0, 0, max(1, len(rows)), len(FILE_INDEX_KEYS) - 1)
    index_sheet.set_column(0, 0, 34)
    index_sheet.set_column(1, len(FILE_INDEX_KEYS) - 1, 30)

    summary = workbook.add_worksheet("汇总")
    summary.hide_gridlines(2)
    summary.set_column("A:A", 28)
    summary.set_column("B:B", 18)
    summary.write("A1", "数据集采集汇总", section)
    summary.write("A3", "样本总数")
    summary.write_formula("B3", f"=COUNTA('标签'!A2:A{len(rows) + 1})", None, len(rows))
    summary.write("A4", "标记有效样本")
    valid_letter = _excel_column(valid_column)
    split_letter = _excel_column(split_column)
    summary.write_formula("B4", f'=COUNTIF(\'标签\'!{valid_letter}2:{valid_letter}{len(rows) + 1},"是")', None, len(rows))
    summary.write("A5", "训练集样本")
    summary.write_formula("B5", f'=COUNTIF(\'标签\'!{split_letter}2:{split_letter}{len(rows) + 1},"train")', None, sum(row.get("dataset_split") == "train" for row in rows))
    summary.write("A6", "监督配对样本")
    summary.write_number("B6", len(supervised_rows))
    summary.write("A7", "实验元数据", section)
    summary.write("A8", json.dumps(metadata, ensure_ascii=False, indent=2), note)
    summary.set_row(7, 120)

    guide = workbook.add_worksheet("字段说明")
    guide.hide_gridlines(2)
    guide.freeze_panes(1, 0)
    guide.write_row(0, 0, ["字段名", "Excel 列名", "说明"], header)
    for row_index, (key, header_text, description) in enumerate(LABEL_COLUMNS, 1):
        guide.write_row(row_index, 0, [key, header_text, description])
    guide.set_column("A:A", 28)
    guide.set_column("B:B", 24)
    guide.set_column("C:C", 60)
    workbook.close()
