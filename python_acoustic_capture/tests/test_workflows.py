from pathlib import Path
import csv
import json

import numpy as np
import pytest
import soundfile as sf

from acoustic_capture.audio import CaptureResult, SimulatedBackend
from acoustic_capture.check import capture_input_check, capture_silent_duplex_check
from acoustic_capture.config import ExperimentConfig, load_config
from acoustic_capture.general import capture_general_io
from acoustic_capture.rir import capture_rir, rir_delay_acceptance
from acoustic_capture.scene import (
    PAIRING_STRATEGY,
    _safe_label,
    _shared_hardware_clock,
    build_paired_sequence,
    capture_scene_block,
    discover_source_pairs,
)
from acoustic_capture.labels import import_reviewed_labels
from acoustic_capture.speech_dataset import compile_speech_dataset
from acoustic_capture.storage import _safe_name


def test_simulated_backend_streams_standalone_recording_to_wav(tmp_path: Path):
    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.audio.sample_rate = 8_000
    cfg.audio.block_size = 80
    cfg.audio.input_channels = [1, 2, 4]
    backend = SimulatedBackend(cfg.audio)
    progress: list[int] = []
    backend.set_progress_callback(
        lambda update: progress.append(int(update.get("frames", 0)))
    )
    output = tmp_path / "standalone" / "three_mics.wav"

    status = backend.record_to_file(
        output,
        stop_requested=lambda: len(progress) >= 3,
        subtype="FLOAT",
    )

    data, sample_rate = sf.read(output, always_2d=True, dtype="float32")
    assert sample_rate == 8_000
    assert data.shape == (240, 3)
    assert status["frames"] == 240
    assert status["channels"] == 3
    assert status["duration_s"] == pytest.approx(0.03)
    assert sf.info(output).subtype == "FLOAT"


def test_rir_end_to_end(tmp_path: Path):
    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.sweep.duration_s = 0.25
    cfg.sweep.pre_silence_s = 0.05
    cfg.sweep.post_silence_s = 0.1
    cfg.sweep.rir_duration_s = 0.1
    cfg.repeats.fixed_count = 5
    cfg.repeats.minimum = 5
    cfg.repeats.maximum = 10
    cfg.repeats.required_stable_takes = 1
    cfg.repeats.correlation_threshold = 0.9
    cfg.repeats.pause_s = 0
    cfg.storage.root = str(tmp_path)
    cfg.storage.compute_sha256 = False
    store = capture_rir(cfg, SimulatedBackend(cfg.audio), log=lambda _: None)
    average, fs = sf.read(store.path("processed/average_rir.wav"), always_2d=True)
    assert fs == cfg.audio.sample_rate
    assert average.shape == (round(cfg.sweep.rir_duration_s * fs), 2)
    # The simulator places microphone 2 seven samples later.  RIR averaging
    # must keep that physical inter-microphone time difference.
    assert int(np.argmax(np.abs(average[:, 1]))) - int(np.argmax(np.abs(average[:, 0]))) == 7
    assert store.path("processed/average_rir_mic_01.wav").is_file()
    assert store.path("processed/average_rir_mic_02.wav").is_file()
    assert store.path("processed/selected_rir.wav").is_file()
    assert store.path("processed/all_accepted_mean_rir.wav").is_file()
    assert store.path("processed/mean_recon.wav").is_file()
    assert store.path("processed/recon_001.wav").is_file()
    assert store.manifest["summary"]["mean_rir_per_microphone"] == [
        "processed/average_rir_mic_01.wav",
        "processed/average_rir_mic_02.wav",
    ]
    assert store.manifest["summary"]["attempted_takes"] == 5
    assert store.manifest["summary"]["accepted_takes"] == [1, 2, 3, 4, 5]
    assert store.manifest["summary"]["selection_method"] == (
        "all_accepted_aligned_mean"
    )
    assert store.manifest["summary"]["selected_take_ids"] == [1, 2, 3, 4, 5]
    accepted_rirs = np.stack(
        [
            sf.read(
                store.path(f"processed/take_{take:03d}_rir.wav"),
                always_2d=True,
                dtype="float32",
            )[0]
            for take in range(1, 6)
        ]
    )
    assert np.allclose(average, np.mean(accepted_rirs, axis=0), atol=1e-7)
    assert store.manifest["summary"]["convergence"]["converged"] is None
    assert store.manifest["summary"]["convergence"]["stop_reason"] == (
        "fixed_attempt_count_completed"
    )
    reconstruction = store.manifest["summary"]["average_rir_reconstruction"]
    assert len(reconstruction["per_take"]) == 5
    assert len(reconstruction["per_channel_summary"]) == 2
    assert all(
        item["self_reconstruction"]["mean_mse"] >= 0
        for item in [
            json.loads(
                store.path(f"metrics/take_{take:03d}.json").read_text(
                    encoding="utf-8"
                )
            )
            for take in range(1, 6)
        ]
    )
    assert store.manifest["status"] == "completed"
    delay_acceptance = store.manifest["summary"]["two_channel_delay_acceptance"]
    assert delay_acceptance["applicable"] is True
    assert "gcc_phat_delay_samples" in delay_acceptance
    assert "low_frequency_group_delay_samples" in delay_acceptance


def test_two_channel_rir_delay_algorithms_agree():
    fs = 48_000
    first = np.zeros(4096, dtype=np.float32)
    first[150] = 1.0
    first[310] = 0.2
    second = np.zeros_like(first)
    second[7:] = first[:-7]

    result = rir_delay_acceptance(np.column_stack((first, second)), fs)

    assert result["passed"] is True
    assert abs(result["gcc_phat_delay_samples"] - 7) < 0.1
    assert abs(result["low_frequency_group_delay_samples"] - 7) < 0.25


def test_legacy_adaptive_rir_config_migrates_to_reconstruct_average(tmp_path: Path):
    path = tmp_path / "legacy.yaml"
    path.write_text(
        "repeats:\n  strategy: adaptive_select\n  fixed_count: 4\n",
        encoding="utf-8",
    )

    config = load_config(path)

    assert config.repeats.strategy == "reconstruct_average"
    assert config.repeats.fixed_count == 4


def test_rir_stop_keeps_completed_takes_as_a_partial_average(tmp_path: Path):
    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.sweep.duration_s = 0.1
    cfg.sweep.pre_silence_s = 0.02
    cfg.sweep.post_silence_s = 0.05
    cfg.sweep.rir_duration_s = 0.05
    cfg.repeats.fixed_count = 4
    cfg.repeats.minimum = 2
    cfg.repeats.maximum = 4
    cfg.repeats.required_stable_takes = 2
    cfg.repeats.pause_s = 0
    cfg.storage.root = str(tmp_path)
    cfg.storage.compute_sha256 = False
    stopped = False

    def stop_after_first(_path, take):
        nonlocal stopped
        stopped = take == 1

    store = capture_rir(
        cfg,
        SimulatedBackend(cfg.audio),
        log=lambda _: None,
        progress=stop_after_first,
        stop_requested=lambda: stopped,
    )

    assert store.manifest["status"] == "cancelled"
    assert store.manifest["summary"]["accepted_takes"] == [1]
    assert store.manifest["summary"]["partial_average"] is True
    assert store.path("processed/average_rir.wav").is_file()


def test_rir_fixed_count_records_exact_attempts_and_defers_selection(tmp_path: Path):
    class FirstTakeXrunBackend(SimulatedBackend):
        def __init__(self, config):
            super().__init__(config)
            self.take = 0

        def play_record(self, output: np.ndarray) -> CaptureResult:
            self.take += 1
            result = super().play_record(output)
            result.status["xrun"] = self.take == 1
            return result

    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.sweep.duration_s = 0.1
    cfg.sweep.pre_silence_s = 0.02
    cfg.sweep.post_silence_s = 0.05
    cfg.sweep.rir_duration_s = 0.05
    cfg.repeats.strategy = "fixed_count"
    cfg.repeats.fixed_count = 3
    # Raw-selection mode still records the exact requested number and keeps a
    # reference mean without choosing a best-single take.
    cfg.repeats.minimum = 5
    cfg.repeats.maximum = 8
    cfg.repeats.required_stable_takes = 1
    cfg.repeats.pause_s = 0
    cfg.storage.root = str(tmp_path)
    cfg.storage.compute_sha256 = False

    store = capture_rir(
        cfg, FirstTakeXrunBackend(cfg.audio), log=lambda _: None
    )
    summary = store.manifest["summary"]

    assert summary["capture_strategy"] == "fixed_count"
    assert summary["requested_fixed_attempts"] == 3
    assert summary["attempted_takes"] == 3
    assert summary["accepted_takes"] == [2, 3]
    assert summary["rejected_takes"] == [1]
    assert summary["convergence"]["stop_reason"] == "fixed_attempt_count_completed"
    assert summary["offline_reselection"]["selection_deferred"] is True
    assert len(summary["offline_reselection"]["takes"]) == 3
    assert all(store.path(f"raw/take_{take:03d}.wav").is_file() for take in (1, 2, 3))
    assert store.path("processed/selected_rir.wav").is_file()
    assert store.manifest["status"] == "completed"


def test_rir_supports_more_than_two_microphones(tmp_path: Path):
    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.audio.input_channels = [1, 2, 3, 4]
    cfg.sweep.duration_s = 0.2
    cfg.sweep.pre_silence_s = 0.05
    cfg.sweep.post_silence_s = 0.1
    cfg.sweep.rir_duration_s = 0.08
    cfg.repeats.fixed_count = 2
    cfg.repeats.minimum = 2
    cfg.repeats.maximum = 2
    cfg.repeats.required_stable_takes = 1
    cfg.repeats.correlation_threshold = 0.85
    cfg.repeats.pause_s = 0
    cfg.storage.root = str(tmp_path)
    cfg.storage.compute_sha256 = False
    cfg.validate()
    store = capture_rir(cfg, SimulatedBackend(cfg.audio), log=lambda _: None)
    average, _ = sf.read(store.path("processed/average_rir.wav"), always_2d=True)
    assert average.shape[1] == 4


def test_rir_rejects_take_with_audio_xrun(tmp_path: Path):
    class FirstTakeXrunBackend(SimulatedBackend):
        def __init__(self, config):
            super().__init__(config)
            self.take = 0

        def play_record(self, output: np.ndarray) -> CaptureResult:
            self.take += 1
            result = super().play_record(output)
            result.status["xrun"] = self.take == 1
            return result

    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.sweep.duration_s = 0.2
    cfg.sweep.pre_silence_s = 0.05
    cfg.sweep.post_silence_s = 0.1
    cfg.sweep.rir_duration_s = 0.08
    cfg.repeats.fixed_count = 3
    cfg.repeats.minimum = 2
    cfg.repeats.maximum = 3
    cfg.repeats.required_stable_takes = 1
    cfg.repeats.correlation_threshold = 0.85
    cfg.repeats.pause_s = 0
    cfg.storage.root = str(tmp_path)
    cfg.storage.compute_sha256 = False
    store = capture_rir(cfg, FirstTakeXrunBackend(cfg.audio), log=lambda _: None)
    assert store.manifest["summary"]["rejected_takes"] == [1]
    assert store.manifest["summary"]["accepted_takes"] == [2, 3]


def test_rir_rejects_silent_microphones_instead_of_averaging_zero_ir(tmp_path: Path):
    class SilentBackend(SimulatedBackend):
        def play_record(self, output: np.ndarray) -> CaptureResult:
            return CaptureResult(
                np.zeros((len(output), len(self.config.input_channels)), dtype=np.float32),
                {"backend": "silent", "xrun": False},
            )

    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.sweep.duration_s = 0.05
    cfg.sweep.pre_silence_s = 0.02
    cfg.sweep.post_silence_s = 0.03
    cfg.sweep.rir_duration_s = 0.03
    cfg.repeats.fixed_count = 1
    cfg.repeats.minimum = 1
    cfg.repeats.maximum = 1
    cfg.repeats.pause_s = 0
    cfg.storage.root = str(tmp_path)
    cfg.storage.compute_sha256 = False

    with np.testing.assert_raises_regex(RuntimeError, "只有 0 次有效"):
        capture_rir(cfg, SilentBackend(cfg.audio), log=lambda _: None)
    run = next(path for path in tmp_path.iterdir() if path.is_dir())
    metrics = json.loads((run / "metrics" / "take_001.json").read_text(encoding="utf-8"))
    assert "扫频信噪比不足" in metrics["rejection_reasons"]


def test_manual_names_are_safe_for_windows_and_keep_unicode_dataset_identity():
    assert _safe_name('耳机 A: 角度 90? / 高度 170*') == "耳机_A-_角度_90-_-_高度_170-"
    assert _safe_name("CON") == "_CON"
    assert _safe_label("耳机A / 角度90°") == "耳机A_角度90"


def test_selectable_scene_block(tmp_path: Path):
    fs = 48_000
    tone = np.sin(2 * np.pi * 440 * np.arange(fs // 10) / fs).astype(np.float32)
    target, interferer = tmp_path / "target.wav", tmp_path / "interferer.wav"
    sf.write(target, tone, fs)
    sf.write(interferer, tone, fs)
    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.scene.items = ["target_only", "mixture"]
    cfg.scene.duration_s = 0.1
    cfg.scene.target_file = str(target)
    cfg.scene.interferer_file = str(interferer)
    cfg.scene.gap_s = 0
    cfg.scene.countdown_s = 0
    cfg.storage.root = str(tmp_path / "runs")
    cfg.storage.compute_sha256 = False
    cfg.metadata["scene_id"] = "hs01_w01_b00_int090_h170"
    class CountingBackend(SimulatedBackend):
        def __init__(self, config):
            super().__init__(config)
            self.play_record_calls = 0

        def play_record(self, output: np.ndarray) -> CaptureResult:
            self.play_record_calls += 1
            return super().play_record(output)

    backend = CountingBackend(cfg.audio)
    store = capture_scene_block(cfg, backend, log=lambda _: None)
    assert backend.play_record_calls == 1
    assert store.path("raw/rep_001_target_only_mics.wav").is_file()
    assert store.path("raw/rep_001_mixture_mics.wav").is_file()
    assert not store.path("raw/rep_001_ambient_mics.wav").exists()
    target_playback, _ = sf.read(
        store.path("references/rep_001_target_only_playback.wav"), always_2d=True
    )
    mixture_playback, _ = sf.read(
        store.path("references/rep_001_mixture_playback.wav"), always_2d=True
    )
    assert target_playback.shape[1] == mixture_playback.shape[1] == 2
    assert np.allclose(target_playback[:, 1], 0)
    assert np.array_equal(target_playback[:, 0], mixture_playback[:, 0])
    with store.path("labels.csv").open(encoding="utf-8-sig", newline="") as handle:
        row = next(csv.DictReader(handle))
    assert row["supervision_ready"] == "是"
    assert row["sample_id"].startswith("hs01_w01_b00_int090_h170__")
    assert row["target_mixture_sample_aligned"] == "是"
    assert row["capture_strategy"] == "single_stream_paired_sequence"
    assert row["paired_sequence_recording"].endswith("paired_sequence_mics.wav")
    assert row["segment_layout"].endswith("paired_sequence_layout.json")
    assert store.path("labels.xlsx").is_file()
    assert store.path("labels.csv").is_file()
    assert store.path("labels.jsonl").is_file()
    assert store.path("raw/labels.csv").is_file()
    assert store.path("raw/labels.jsonl").is_file()
    assert store.path("raw/supervised_pairs.csv").is_file()
    assert store.path("raw/supervised_pairs.jsonl").is_file()


def test_target_and_mixed_pair_does_not_require_interferer_only(tmp_path: Path):
    fs = 16_000
    tone = np.sin(2 * np.pi * 440 * np.arange(800) / fs).astype(np.float32)
    target, interferer = tmp_path / "target.wav", tmp_path / "interferer.wav"
    sf.write(target, tone, fs)
    sf.write(interferer, tone, fs)
    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.audio.sample_rate = fs
    cfg.scene.items = ["target_only", "mixture"]
    cfg.scene.duration_s = 0.05
    cfg.scene.target_file = str(target)
    cfg.scene.interferer_file = str(interferer)
    cfg.scene.countdown_s = 0
    cfg.scene.gap_s = 0
    cfg.storage.root = str(tmp_path / "runs")
    cfg.storage.compute_sha256 = False

    store = capture_scene_block(cfg, SimulatedBackend(cfg.audio), log=lambda _: None)
    assert store.path("raw/rep_001_target_only_mics.wav").is_file()
    assert store.path("raw/rep_001_mixture_mics.wav").is_file()
    assert not store.path("raw/rep_001_interferer_only_mics.wav").exists()
    with store.path("labels.csv").open(encoding="utf-8-sig", newline="") as handle:
        row = next(csv.DictReader(handle))
    assert row["capture_type"] == "supervised_pair"
    assert row["target_mixture_sample_aligned"] == "是"
    assert row["interferer_mixture_sample_aligned"] == "否"
    assert row["supervision_ready"] == "是"


def test_scene_rejects_one_silent_microphone_for_supervision(tmp_path: Path):
    class OneSilentChannelBackend(SimulatedBackend):
        def play_record(self, output: np.ndarray) -> CaptureResult:
            result = super().play_record(output)
            result.microphones[:, 1] = 0
            return result

    fs = 16_000
    signal = np.sin(2 * np.pi * 440 * np.arange(800) / fs).astype(np.float32)
    target, interferer = tmp_path / "target.wav", tmp_path / "interferer.wav"
    sf.write(target, signal, fs)
    sf.write(interferer, signal, fs)
    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.audio.sample_rate = fs
    cfg.scene.items = ["target_only", "mixture"]
    cfg.scene.target_file = str(target)
    cfg.scene.interferer_file = str(interferer)
    cfg.scene.duration_s = 0.05
    cfg.scene.countdown_s = 0
    cfg.scene.gap_s = 0
    cfg.storage.root = str(tmp_path / "runs")
    cfg.storage.compute_sha256 = False

    store = capture_scene_block(cfg, OneSilentChannelBackend(cfg.audio), log=lambda _: None)
    with store.path("labels.csv").open(encoding="utf-8-sig", newline="") as handle:
        row = next(csv.DictReader(handle))
    assert row["quality_flag"] == "存在静音通道"
    assert row["supervision_ready"] == "否"


def test_scene_failure_checkpoints_and_resume_skips_completed_pair(tmp_path: Path):
    fs = 16_000
    target_folder, interferer_folder = tmp_path / "targets", tmp_path / "interferers"
    target_folder.mkdir()
    interferer_folder.mkdir()
    for index in range(3):
        signal = np.sin(2 * np.pi * (300 + index * 100) * np.arange(320) / fs)
        sf.write(target_folder / f"target_{index}.wav", signal, fs)
        sf.write(interferer_folder / f"noise_{index}.wav", signal, fs)
    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.audio.sample_rate = fs
    cfg.scene.source_mode = "folders"
    cfg.scene.target_folder = str(target_folder)
    cfg.scene.interferer_folder = str(interferer_folder)
    cfg.scene.items = ["target_only", "mixture"]
    cfg.scene.duration_s = 0.02
    cfg.scene.countdown_s = 0
    cfg.scene.gap_s = 0
    cfg.storage.root = str(tmp_path / "runs")
    cfg.storage.compute_sha256 = False

    class FailSecondBackend(SimulatedBackend):
        def __init__(self, audio):
            super().__init__(audio)
            self.calls = 0

        def play_record(self, output: np.ndarray) -> CaptureResult:
            self.calls += 1
            if self.calls == 2:
                raise RuntimeError("injected failure")
            return super().play_record(output)

    with pytest.raises(RuntimeError, match="injected failure"):
        capture_scene_block(cfg, FailSecondBackend(cfg.audio), log=lambda _: None)
    run = next((tmp_path / "runs").iterdir())
    manifest = json.loads((run / "manifest.json").read_text(encoding="utf-8"))
    assert manifest["status"] == "failed"
    assert len((run / "labels.jsonl").read_text(encoding="utf-8").splitlines()) == 1
    cfg.scene.resume_run = str(run)
    resumed_backend = FailSecondBackend(cfg.audio)
    resumed_backend.calls = -100  # disable the injected second-call failure
    store = capture_scene_block(cfg, resumed_backend, log=lambda _: None)
    assert store.root == run
    assert store.manifest["status"] == "completed"
    assert len((run / "labels.jsonl").read_text(encoding="utf-8").splitlines()) == 3
    assert resumed_backend.calls == -98


def test_scene_resume_rejects_changed_level_or_physical_labels(tmp_path: Path):
    fs = 16_000
    signal = np.sin(2 * np.pi * 440 * np.arange(320) / fs)
    target, interferer = tmp_path / "target.wav", tmp_path / "interferer.wav"
    sf.write(target, signal, fs)
    sf.write(interferer, signal, fs)
    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.audio.sample_rate = fs
    cfg.scene.items = ["target_only", "mixture"]
    cfg.scene.target_file = str(target)
    cfg.scene.interferer_file = str(interferer)
    cfg.scene.duration_s = 0.02
    cfg.scene.countdown_s = 0
    cfg.scene.gap_s = 0
    cfg.storage.root = str(tmp_path / "runs")
    cfg.storage.compute_sha256 = False
    cfg.metadata = {"wearing_id": "w01"}

    class AlwaysFail(SimulatedBackend):
        def play_record(self, output: np.ndarray) -> CaptureResult:
            raise RuntimeError("stop here")

    with pytest.raises(RuntimeError, match="stop here"):
        capture_scene_block(cfg, AlwaysFail(cfg.audio), log=lambda _: None)
    run = next((tmp_path / "runs").iterdir())
    cfg.scene.resume_run = str(run)
    cfg.scene.target_level_dbfs = -12
    cfg.metadata["wearing_id"] = "w02"
    with pytest.raises(ValueError, match="不能续采"):
        capture_scene_block(cfg, SimulatedBackend(cfg.audio), log=lambda _: None)


def test_source_indexes_and_reviewed_excel_feed_training_labels(tmp_path: Path):
    fs = 16_000
    target_folder, interferer_folder = tmp_path / "targets", tmp_path / "interferers"
    target_folder.mkdir()
    interferer_folder.mkdir()
    signal = np.sin(2 * np.pi * 440 * np.arange(320) / fs)
    sf.write(target_folder / "utterance.wav", signal, fs)
    sf.write(interferer_folder / "noise.wav", signal, fs)
    target_index = tmp_path / "target_index.csv"
    target_index.write_text(
        "relative_path,speaker_id,utterance_id\nutterance.wav,spk007,utt042\n",
        encoding="utf-8-sig",
    )
    interferer_index = tmp_path / "interferer_index.csv"
    interferer_index.write_text(
        "relative_path,noise_id,noise_class\nnoise.wav,n008,cafe\n",
        encoding="utf-8-sig",
    )
    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.audio.sample_rate = fs
    cfg.scene.source_mode = "folders"
    cfg.scene.target_folder = str(target_folder)
    cfg.scene.interferer_folder = str(interferer_folder)
    cfg.scene.target_index_csv = str(target_index)
    cfg.scene.interferer_index_csv = str(interferer_index)
    cfg.scene.items = ["target_only", "mixture"]
    cfg.scene.duration_s = 0.02
    cfg.scene.countdown_s = 0
    cfg.scene.gap_s = 0
    cfg.storage.root = str(tmp_path / "runs")
    cfg.storage.compute_sha256 = False
    store = capture_scene_block(cfg, SimulatedBackend(cfg.audio), log=lambda _: None)

    with store.path("labels.csv").open(encoding="utf-8-sig", newline="") as handle:
        row = next(csv.DictReader(handle))
    assert (row["speaker_id"], row["utterance_id"]) == ("spk007", "utt042")
    assert (row["noise_id"], row["noise_class"]) == ("n008", "cafe")
    from openpyxl import load_workbook

    workbook = load_workbook(store.path("labels.xlsx"))
    sheet = workbook["标签"]
    headers = [cell.value for cell in sheet[1]]
    sheet.cell(2, headers.index("人工标签") + 1, "人工通过")
    sheet.cell(2, headers.index("数据集划分") + 1, "test")
    workbook.save(store.path("labels.xlsx"))
    import_reviewed_labels(store.root)
    dataset = compile_speech_dataset(tmp_path / "runs", tmp_path / "dataset", copy_audio=False)
    reviewed = json.loads(
        (dataset / "indexes" / "speech_samples.jsonl").read_text(encoding="utf-8").splitlines()[0]
    )
    assert reviewed["manual_label"] == "人工通过"
    assert reviewed["dataset_split"] == "test"


def test_paired_sequence_uses_identical_target_samples_and_shared_boundaries():
    cfg = ExperimentConfig()
    cfg.audio.sample_rate = 1_000
    cfg.scene.items = ["target_only", "interferer_only", "mixture"]
    cfg.scene.gap_s = 0.01
    target = np.linspace(-0.5, 0.5, 100, dtype=np.float32)
    interferer = np.linspace(0.25, -0.25, 100, dtype=np.float32)

    sequence, segments = build_paired_sequence(target, interferer, cfg)

    assert len(sequence) == 10 + 3 * (100 + 10)
    target_only = segments["target_only"]["playback"]
    interferer_only = segments["interferer_only"]["playback"]
    mixture = segments["mixture"]["playback"]
    assert np.array_equal(target_only[:, 0], mixture[:, 0])
    assert np.array_equal(interferer_only[:, 1], mixture[:, 1])
    assert all(segment["sample_count"] == 100 for segment in segments.values())


def test_supervised_mixture_requires_target_only():
    cfg = ExperimentConfig()
    cfg.scene.items = ["interferer_only", "mixture"]
    with np.testing.assert_raises_regex(ValueError, "requires target_only"):
        cfg.validate()


def test_supervision_requires_one_verified_duplex_device_clock():
    cfg = ExperimentConfig()
    cfg.audio.backend = "sounddevice"
    cfg.audio.input_device = "RME input"
    cfg.audio.output_device = "Beosound output"
    assert not _shared_hardware_clock(cfg)
    cfg.audio.input_device = cfg.audio.output_device = "ASIO Fireface USB"
    assert _shared_hardware_clock(cfg)


def test_pure_target_and_pure_interferer_scenes_are_supported(tmp_path: Path):
    fs = 16_000
    tone = np.sin(2 * np.pi * 330 * np.arange(800) / fs).astype(np.float32)
    target = tmp_path / "target.wav"
    interferer = tmp_path / "interferer.wav"
    sf.write(target, tone, fs)
    sf.write(interferer, tone, fs)

    for item, expected, absent in (
        ("target_only", "target_only", "interferer_only"),
        ("interferer_only", "interferer_only", "target_only"),
    ):
        cfg = ExperimentConfig()
        cfg.audio.backend = "simulated"
        cfg.audio.sample_rate = fs
        cfg.scene.items = [item]
        cfg.scene.duration_s = 0.05
        cfg.scene.target_file = str(target)
        cfg.scene.interferer_file = str(interferer)
        cfg.scene.countdown_s = 0
        cfg.scene.gap_s = 0
        cfg.storage.root = str(tmp_path / item)
        cfg.storage.compute_sha256 = False
        store = capture_scene_block(cfg, SimulatedBackend(cfg.audio), log=lambda _: None)

        assert store.path(f"raw/rep_001_{expected}_mics.wav").is_file()
        assert not store.path(f"raw/rep_001_{absent}_mics.wav").exists()
        with store.path("labels.csv").open(encoding="utf-8-sig", newline="") as handle:
            row = next(csv.DictReader(handle))
        assert row["supervision_ready"] == "否"


def test_folder_scene_batch_cycles_files_and_writes_labels(tmp_path: Path):
    fs = 16_000
    target_folder = tmp_path / "targets"
    interferer_folder = tmp_path / "interferers"
    target_folder.mkdir()
    interferer_folder.mkdir()
    time_axis = np.arange(fs // 20) / fs
    for index, frequency in enumerate((220, 330, 440), 1):
        sf.write(target_folder / f"speaker_{index}.wav", np.sin(2 * np.pi * frequency * time_axis), fs)
    for index, frequency in enumerate((700, 900), 1):
        sf.write(interferer_folder / f"noise_{index}.wav", np.sin(2 * np.pi * frequency * time_axis), fs)

    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.audio.sample_rate = fs
    cfg.scene.source_mode = "folders"
    cfg.scene.target_folder = str(target_folder)
    cfg.scene.interferer_folder = str(interferer_folder)
    cfg.scene.pairing_seed = 123
    cfg.scene.duration_s = 0.05
    cfg.scene.ambient_duration_s = 0.05
    cfg.scene.items = ["target_only", "interferer_only", "mixture"]
    cfg.scene.countdown_s = 0
    cfg.scene.gap_s = 0
    cfg.scene.label_prefix = "batch"
    cfg.storage.root = str(tmp_path / "runs")
    cfg.storage.compute_sha256 = False
    planned_pairs = discover_source_pairs(cfg.scene)
    store = capture_scene_block(cfg, SimulatedBackend(cfg.audio), log=lambda _: None)

    with store.path("labels.csv").open(encoding="utf-8-sig", newline="") as handle:
        labels = list(csv.DictReader(handle))
    assert len(labels) == 3
    assert [Path(row["target_source"]).name for row in labels] == [
        "speaker_1.wav",
        "speaker_2.wav",
        "speaker_3.wav",
    ]
    assert [Path(row["interferer_source"]).name for row in labels] == [
        pair.interferer.name for pair in planned_pairs
    ]
    assert all(row["pairing_seed"] == "123" for row in labels)
    assert all(row["pairing_strategy"] == PAIRING_STRATEGY for row in labels)
    assert all(row["automatic_label"].startswith("batch_") for row in labels)
    assert all(row["duration_s"] == "0.05" for row in labels)
    assert store.manifest["summary"]["label_rows"] == 3


def test_folder_sources_are_hashed_lazily_per_captured_pair(tmp_path: Path, monkeypatch):
    fs = 16_000
    target_folder = tmp_path / "targets"
    interferer_folder = tmp_path / "interferers"
    target_folder.mkdir()
    interferer_folder.mkdir()
    signal = np.zeros(160, dtype=np.float32)
    for index in range(4):
        sf.write(target_folder / f"target_{index}.wav", signal, fs)
        sf.write(interferer_folder / f"noise_{index}.wav", signal, fs)

    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.audio.sample_rate = fs
    cfg.sweep.end_hz = 7_000
    cfg.scene.source_mode = "folders"
    cfg.scene.target_folder = str(target_folder)
    cfg.scene.interferer_folder = str(interferer_folder)
    cfg.scene.items = ["target_only", "mixture"]
    cfg.scene.duration_s = 0.01
    cfg.scene.countdown_s = 0
    cfg.scene.gap_s = 0
    cfg.storage.root = str(tmp_path / "runs")
    calls = []
    monkeypatch.setattr("acoustic_capture.scene.sha256", lambda path: calls.append(Path(path)) or "hash")
    stop = False

    def progress(update):
        nonlocal stop
        if update.get("event") == "pair_prepared":
            stop = True

    store = capture_scene_block(
        cfg,
        SimulatedBackend(cfg.audio),
        log=lambda _: None,
        stop_requested=lambda: stop,
        progress=progress,
    )
    assert store.manifest["status"] == "cancelled"
    assert len(calls) == 2


def test_seeded_pairing_is_linear_reproducible_and_changes_with_seed(tmp_path: Path, monkeypatch):
    cfg = ExperimentConfig().scene
    cfg.source_mode = "folders"
    cfg.target_folder = str(tmp_path / "targets")
    cfg.interferer_folder = str(tmp_path / "interferers")
    fake_targets = [tmp_path / f"t{index}.wav" for index in range(2_000)]
    fake_interferers = [tmp_path / f"n{index}.wav" for index in range(2_000)]

    def plan(seed: int):
        cfg.pairing_seed = seed
        scans = iter((fake_targets, fake_interferers))
        monkeypatch.setattr(
            "acoustic_capture.scene._scan_audio_folder", lambda *_: next(scans)
        )
        return discover_source_pairs(cfg)

    first = plan(17)
    repeated = plan(17)
    changed = plan(18)
    assert len(first) == 2_000
    assert [pair.target for pair in first] == fake_targets
    assert [pair.interferer for pair in first] == [
        pair.interferer for pair in repeated
    ]
    assert [pair.interferer for pair in first] != [
        pair.interferer for pair in changed
    ]


def test_speech_dataset_packages_multichannel_pairs_and_flattened_labels(tmp_path: Path):
    fs = 16_000
    time_axis = np.arange(800) / fs
    target = tmp_path / "target.wav"
    interferer = tmp_path / "interferer.wav"
    sf.write(target, np.sin(2 * np.pi * 330 * time_axis), fs)
    sf.write(interferer, np.sin(2 * np.pi * 770 * time_axis), fs)

    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.audio.sample_rate = fs
    cfg.audio.input_channels = [1, 2, 3]
    cfg.scene.items = ["target_only", "interferer_only", "mixture"]
    cfg.scene.target_file = str(target)
    cfg.scene.interferer_file = str(interferer)
    cfg.scene.duration_s = 0.05
    cfg.scene.countdown_s = 0
    cfg.scene.gap_s = 0
    cfg.storage.root = str(tmp_path / "runs")
    cfg.storage.compute_sha256 = False
    cfg.metadata = {
        "project_id": "speech_v1",
        "scene_id": "head02_hs01_w03_int090_h170",
        "experiment_id": "head02_hs01_w03_int090_h170",
        "room_id": "lab_a",
        "artificial_head_id": "head02",
        "headset_model_id": "model_a",
        "headset_unit_id": "hs01",
        "wearing_id": "w03",
        "boom_pose_id": "b00_nominal",
        "microphone_1": "left",
        "microphone_2": "right",
        "microphone_3": "reference",
        "target": {
            "source_id": "mouth01",
            "position_id": "mouth_fixed",
            "azimuth_deg": 0,
            "elevation_deg": 0,
            "height_m": 1.4,
            "distance_m": 0.05,
        },
        "interferer": {
            "source_id": "speaker01",
            "position_id": "az090_h170",
            "azimuth_deg": 90,
            "elevation_deg": 17,
            "height_m": 1.7,
            "distance_m": 1.0,
        },
    }
    store = capture_scene_block(cfg, SimulatedBackend(cfg.audio), log=lambda _: None)

    with store.path("labels.csv").open(encoding="utf-8-sig", newline="") as handle:
        label = next(csv.DictReader(handle))
    assert label["capture_type"] == "supervised_pair"
    assert label["supervision_pair_id"] == label["sample_id"]
    assert label["artificial_head_id"] == "head02"
    assert label["interferer_azimuth_deg"] == "90"
    assert label["microphone_channels"] == "1,2,3"
    assert store.path("supervised_pairs.csv").is_file()

    dataset = compile_speech_dataset(tmp_path / "runs", tmp_path / "speech_dataset")
    manifest = json.loads((dataset / "dataset_manifest.json").read_text(encoding="utf-8"))
    with (dataset / "indexes" / "supervised_pairs.csv").open(
        encoding="utf-8-sig", newline=""
    ) as handle:
        pair = next(csv.DictReader(handle))
    packaged_mixture, _ = sf.read(dataset / pair["mixture_recording"], always_2d=True)
    packaged_target, _ = sf.read(dataset / pair["target_recording"], always_2d=True)
    assert manifest["supervised_pair_count"] == 1
    assert pair["dataset_supervision_ready"] == "是"
    assert pair["mixture_recording"].startswith("audio/target-mixed/")
    assert pair["target_recording"].startswith("audio/target-only/")
    assert pair["interferer_recording"].startswith("audio/interferer-only/")
    assert packaged_mixture.shape == packaged_target.shape == (800, 3)
    assert (dataset / "indexes" / "speech_dataset.xlsx").is_file()


def test_basic_io_actions(tmp_path: Path):
    fs = 48_000
    source = tmp_path / "source.wav"
    sf.write(source, np.zeros(fs // 20, dtype=np.float32), fs)
    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.general.source_file = str(source)
    cfg.general.duration_s = 0.05
    cfg.storage.root = str(tmp_path / "runs")
    cfg.storage.compute_sha256 = False
    for action in ("play", "record", "play_record"):
        cfg.general.action = action
        store = capture_general_io(cfg, SimulatedBackend(cfg.audio), log=lambda _: None)
        assert store.manifest["status"] == "completed"
        assert store.path("references/played.wav").exists() is (action != "record")
        assert store.path("raw/recording.wav").exists() is (action != "play")


def test_basic_io_warns_when_recording_is_all_zero(tmp_path: Path):
    class SilentBackend(SimulatedBackend):
        def record(self, frames: int) -> CaptureResult:
            return CaptureResult(np.zeros((frames, 2), dtype=np.float32), {"backend": "silent-test"})

    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.general.action = "record"
    cfg.general.duration_s = 0.01
    cfg.storage.root = str(tmp_path)
    cfg.storage.compute_sha256 = False
    store = capture_general_io(cfg, SilentBackend(cfg.audio), log=lambda _: None)
    assert store.manifest["summary"]["warnings"]


def test_input_check_warns_when_all_microphones_are_zero(tmp_path: Path):
    class SilentBackend(SimulatedBackend):
        def record(self, frames: int) -> CaptureResult:
            return CaptureResult(np.zeros((frames, 2), dtype=np.float32), {"backend": "silent-test"})

    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.storage.root = str(tmp_path)
    cfg.storage.compute_sha256 = False
    store = capture_input_check(cfg, SilentBackend(cfg.audio), duration_s=0.01)
    assert store.manifest["summary"]["warnings"]


def test_silent_duplex_check_opens_both_output_channels(tmp_path: Path):
    class InspectBackend(SimulatedBackend):
        def play_record(self, output: np.ndarray) -> CaptureResult:
            assert output.shape == (480, 2)
            assert np.count_nonzero(output) == 0
            return CaptureResult(np.zeros((480, 2), dtype=np.float32), {"xrun": False})

    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.storage.root = str(tmp_path)
    cfg.storage.compute_sha256 = False
    store = capture_silent_duplex_check(cfg, InspectBackend(cfg.audio), duration_s=0.01)
    assert store.manifest["status"] == "completed"
    assert store.manifest["summary"]["output_channels_opened"] == 2


def test_silent_duplex_check_reports_audio_xrun(tmp_path: Path):
    class XrunBackend(SimulatedBackend):
        def play_record(self, output: np.ndarray) -> CaptureResult:
            return CaptureResult(np.zeros((len(output), 2), dtype=np.float32), {"xrun": True})

    cfg = ExperimentConfig()
    cfg.audio.backend = "simulated"
    cfg.storage.root = str(tmp_path)
    cfg.storage.compute_sha256 = False
    store = capture_silent_duplex_check(cfg, XrunBackend(cfg.audio), duration_s=0.01)
    assert store.manifest["summary"]["warnings"]
